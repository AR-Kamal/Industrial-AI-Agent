"""Controlled bulk chunk review import/export services."""

import csv
import hashlib
import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .corrections import (
    SplitSegment,
    apply_split,
    content_hash,
    validate_segments,
)
from .models import (
    ChunkMetadataCorrection,
    DocumentChunk,
    KnowledgeDocument,
)

REVIEW_COLUMNS = (
    "chunk_id",
    "source_content_hash",
    "document_id",
    "document_version",
    "sequence",
    "chapter",
    "section",
    "page_start",
    "page_end",
    "token_count",
    "content",
    "contains_warning",
    "contains_caution",
    "processing_warnings",
    "origin",
    "review_status",
    "retrieval_enabled",
    "reviewer_notes",
    "proposed_action",
    "correction_payload",
    "validation_result",
)
ACTIONS = {"NO_CHANGE", "APPROVE", "EXCLUDE", "CORRECT_METADATA", "SPLIT"}
MAX_REVIEW_FILE_BYTES = 20 * 1024 * 1024
MAX_REVIEW_ROWS = 10_000
XML_CONTROL_PATTERN = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


@dataclass(frozen=True)
class ReviewRow:
    row_number: int
    chunk_id: str
    source_content_hash: str
    document_id: str
    document_version: str
    sequence: str
    chapter: str
    section: str
    page_start: int | None
    page_end: int | None
    token_count: int
    content: str
    contains_warning: bool
    contains_caution: bool
    processing_warnings: str
    origin: str
    review_status: str
    retrieval_enabled: bool
    reviewer_notes: str
    proposed_action: str
    correction_payload: dict[str, Any]
    validation_result: str


@dataclass
class RowPlan:
    row: ReviewRow
    chunk: DocumentChunk | None
    errors: list[str]
    summary: str

    def as_report(self) -> dict[str, Any]:
        return {
            "row": self.row.row_number,
            "chunk_id": self.row.chunk_id,
            "action": self.row.proposed_action,
            "valid": not self.errors,
            "errors": self.errors,
            "planned_change": self.summary,
        }


def _spreadsheet_text(value: str) -> str:
    return f"'{value}" if value.startswith(("=", "+", "-", "@")) else value


def _restore_spreadsheet_text(value: Any) -> str:
    text = "" if value is None else str(value)
    if len(text) > 1 and text[0] == "'" and text[1] in "=+-@":
        return text[1:]
    return text


def _review_text(value: str) -> str:
    return XML_CONTROL_PATTERN.sub("", value)


def _canonical_content_line_endings(value: str) -> str:
    """Canonicalize only newline encodings for immutable content comparison."""
    return value.replace("\r\n", "\n").replace("\r", "\n")


def _immutable_content_matches(workbook_content: str, source_content: str) -> bool:
    return _canonical_content_line_endings(
        workbook_content
    ) == _canonical_content_line_endings(_review_text(source_content))


def chunk_export_row(chunk: DocumentChunk) -> dict[str, Any]:
    return {
        "chunk_id": chunk.chunk_id,
        "source_content_hash": chunk.content_hash,
        "document_id": chunk.document_id,
        "document_version": chunk.document_version_id,
        "sequence": str(chunk.sequence),
        "chapter": chunk.chapter,
        "section": chunk.section,
        "page_start": chunk.page_start,
        "page_end": chunk.page_end,
        "token_count": chunk.token_count,
        "content": _review_text(chunk.content),
        "contains_warning": chunk.contains_warning,
        "contains_caution": chunk.contains_caution,
        "processing_warnings": json.dumps(
            chunk.processing_warnings,
            ensure_ascii=False,
        ),
        "origin": chunk.origin,
        "review_status": chunk.review_status,
        "retrieval_enabled": chunk.retrieval_enabled,
        "reviewer_notes": chunk.reviewer_notes,
        "proposed_action": "NO_CHANGE",
        "correction_payload": "",
        "validation_result": "",
    }


def export_review_workbook(
    document_id: str,
    output_path: Path,
) -> int:
    if not KnowledgeDocument.objects.filter(pk=document_id).exists():
        raise ValidationError("Knowledge document not found.")
    chunks = list(
        DocumentChunk.objects.filter(
            document_id=document_id,
            is_current_generation=True,
        )
        .select_related("document_version")
        .order_by("sequence")
    )
    rows = [chunk_export_row(chunk) for chunk in chunks]
    suffix = output_path.suffix.casefold()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if suffix == ".xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Chunk Review"
        sheet.append(list(REVIEW_COLUMNS))
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:U{max(1, len(rows) + 1)}"
        for row in rows:
            sheet.append([row[column] for column in REVIEW_COLUMNS])
            for cell in sheet[sheet.max_row]:
                if isinstance(cell.value, str):
                    cell.data_type = "s"
        workbook.save(output_path)
    elif suffix == ".csv":
        with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=REVIEW_COLUMNS)
            writer.writeheader()
            for row in rows:
                writer.writerow(
                    {
                        key: (
                            _spreadsheet_text(value)
                            if isinstance(value, str)
                            else value
                        )
                        for key, value in row.items()
                    }
                )
    elif suffix == ".json":
        output_path.write_text(
            json.dumps({"rows": rows}, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
    else:
        raise ValidationError("Output must use .xlsx, .csv, or .json.")
    return len(rows)


def _raw_rows(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        raise ValidationError("Review file not found.")
    if path.stat().st_size > MAX_REVIEW_FILE_BYTES:
        raise ValidationError("Review file exceeds the 20 MB limit.")
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
    elif suffix == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        payload_rows = payload.get("rows") if isinstance(payload, dict) else payload
        if not isinstance(payload_rows, list) or not all(
            isinstance(item, dict) for item in payload_rows
        ):
            raise ValidationError("JSON must contain a row list.")
        rows = payload_rows
    elif suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=False)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=False))
        if not values:
            rows = []
        else:
            headers = [str(cell.value or "").strip() for cell in values[0]]
            rows = []
            for cells in values[1:]:
                if any(cell.data_type == "f" for cell in cells):
                    raise ValidationError("Spreadsheet formulas are not allowed.")
                rows.append(
                    {
                        header: cell.value
                        for header, cell in zip(headers, cells, strict=False)
                    }
                )
        workbook.close()
    else:
        raise ValidationError("Review file must use .xlsx, .csv, or .json.")
    if len(rows) > MAX_REVIEW_ROWS:
        raise ValidationError("Review file exceeds the 10,000-row limit.")
    return rows


def _boolean(value: Any, field: str) -> bool:
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().casefold()
    if normalized in {"true", "1", "yes"}:
        return True
    if normalized in {"false", "0", "no", ""}:
        return False
    raise ValidationError(f"{field} must be true or false.")


def _integer(value: Any, field: str, *, optional: bool = False) -> int | None:
    if optional and (value is None or str(value).strip() == ""):
        return None
    try:
        number = int(value)
    except (TypeError, ValueError) as exc:
        raise ValidationError(f"{field} must be an integer.") from exc
    if number < 0:
        raise ValidationError(f"{field} cannot be negative.")
    return number


def _payload(value: Any) -> dict[str, Any]:
    if value is None or str(value).strip() == "":
        return {}
    if isinstance(value, dict):
        return value
    try:
        parsed = json.loads(str(value))
    except json.JSONDecodeError as exc:
        raise ValidationError("correction_payload must be valid JSON.") from exc
    if not isinstance(parsed, dict):
        raise ValidationError("correction_payload must be a JSON object.")
    return parsed


def load_review_rows(path: Path) -> list[ReviewRow]:
    parsed: list[ReviewRow] = []
    raw_rows = _raw_rows(path)
    for row_number, raw in enumerate(raw_rows, start=2):
        missing = [column for column in REVIEW_COLUMNS if column not in raw]
        if missing:
            raise ValidationError(
                f"Row {row_number}: missing columns: {', '.join(missing)}."
            )
        parsed.append(
            ReviewRow(
                row_number=row_number,
                chunk_id=_restore_spreadsheet_text(raw["chunk_id"]).strip(),
                source_content_hash=str(raw["source_content_hash"] or "").strip(),
                document_id=str(raw["document_id"] or "").strip(),
                document_version=str(raw["document_version"] or "").strip(),
                sequence=str(raw["sequence"] or "").strip(),
                chapter=_restore_spreadsheet_text(raw["chapter"]).strip(),
                section=_restore_spreadsheet_text(raw["section"]).strip(),
                page_start=_integer(raw["page_start"], "page_start", optional=True),
                page_end=_integer(raw["page_end"], "page_end", optional=True),
                token_count=int(_integer(raw["token_count"], "token_count") or 0),
                content=_restore_spreadsheet_text(raw["content"]),
                contains_warning=_boolean(
                    raw["contains_warning"],
                    "contains_warning",
                ),
                contains_caution=_boolean(
                    raw["contains_caution"],
                    "contains_caution",
                ),
                processing_warnings=str(raw["processing_warnings"] or ""),
                origin=str(raw["origin"] or "").strip(),
                review_status=str(raw["review_status"] or "").strip(),
                retrieval_enabled=_boolean(
                    raw["retrieval_enabled"],
                    "retrieval_enabled",
                ),
                reviewer_notes=_restore_spreadsheet_text(raw["reviewer_notes"]).strip(),
                proposed_action=str(raw["proposed_action"] or "").strip().upper(),
                correction_payload=_payload(raw["correction_payload"]),
                validation_result=str(raw["validation_result"] or ""),
            )
        )
    return parsed


def resolve_reviewer(username: str | None) -> Any:
    user_model = get_user_model()
    if username:
        reviewer = user_model.objects.filter(username=username).first()
    else:
        candidates = list(
            user_model.objects.filter(is_active=True, is_superuser=True)[:2]
        )
        reviewer = candidates[0] if len(candidates) == 1 else None
    if reviewer is None:
        raise ValidationError(
            "Specify --reviewer with an existing active staff username."
        )
    required = (
        reviewer.is_active
        and reviewer.is_staff
        and reviewer.has_perm("knowledge_base.change_documentchunk")
        and reviewer.has_perm("knowledge_base.add_chunksplitcorrection")
        and reviewer.has_perm("knowledge_base.add_chunkmetadatacorrection")
    )
    if not required:
        raise ValidationError(
            "Reviewer lacks staff chunk-review and correction permissions."
        )
    return reviewer


def _split_segments(row: ReviewRow) -> list[SplitSegment]:
    children = row.correction_payload.get("children")
    if not isinstance(children, list):
        raise ValidationError("SPLIT requires a children array.")
    segments: list[SplitSegment] = []
    for child in children:
        if not isinstance(child, dict):
            raise ValidationError("Every split child must be a JSON object.")
        try:
            segments.append(
                SplitSegment(
                    content=str(child.get("content", "")),
                    chapter=str(child.get("chapter", "")),
                    section=str(child.get("section", "")),
                    page_start=_integer(
                        child.get("page_start"),
                        "child page_start",
                        optional=True,
                    ),
                    page_end=_integer(
                        child.get("page_end"),
                        "child page_end",
                        optional=True,
                    ),
                    contains_warning=_boolean(
                        child.get("contains_warning", False),
                        "child contains_warning",
                    ),
                    contains_caution=_boolean(
                        child.get("contains_caution", False),
                        "child contains_caution",
                    ),
                    retrieval_enabled=_boolean(
                        child.get("retrieval_enabled", False),
                        "child retrieval_enabled",
                    ),
                    reviewer_notes=str(child.get("reviewer_notes", "")).strip(),
                )
            )
        except (TypeError, ValueError) as exc:
            raise ValidationError("Invalid split child data.") from exc
    return segments


def _validate_page_range(row: ReviewRow, chunk: DocumentChunk) -> None:
    if row.page_start is None or row.page_end is None:
        raise ValidationError("A complete page range is required.")
    if row.page_start > row.page_end:
        raise ValidationError("Page start cannot exceed page end.")
    if (
        chunk.page_start is None
        or chunk.page_end is None
        or row.page_start < chunk.page_start
        or row.page_end > chunk.page_end
    ):
        raise ValidationError("Page range must remain inside the source chunk.")


def validate_review_rows(rows: list[ReviewRow]) -> list[RowPlan]:
    ids = [row.chunk_id for row in rows]
    chunks = {
        chunk.chunk_id: chunk
        for chunk in DocumentChunk.objects.filter(chunk_id__in=ids).select_related(
            "document",
            "document_version",
        )
    }
    plans: list[RowPlan] = []
    seen_ids: set[str] = set()
    active_hashes = {
        (chunk.document_version_id, chunk.content_hash): chunk.chunk_id
        for chunk in DocumentChunk.objects.filter(
            is_current_generation=True,
            retrieval_enabled=True,
        )
    }
    for row in rows:
        errors: list[str] = []
        chunk = chunks.get(row.chunk_id)
        action = row.proposed_action
        if row.chunk_id in seen_ids:
            errors.append("Chunk ID appears more than once in the review file.")
        seen_ids.add(row.chunk_id)
        if action not in ACTIONS:
            errors.append("Unsupported proposed_action.")
        if chunk is None:
            errors.append("Unknown chunk ID.")
        else:
            if row.source_content_hash != chunk.content_hash:
                errors.append("Source content hash mismatch.")
            if not _immutable_content_matches(row.content, chunk.content):
                errors.append("Exported chunk content must not be edited.")
            if (
                row.document_id != chunk.document_id
                or row.document_version != chunk.document_version_id
            ):
                errors.append("Document provenance mismatch.")
            if action != "NO_CHANGE" and (
                not chunk.is_current_generation
                or chunk.review_status == DocumentChunk.ReviewStatus.SUPERSEDED
            ):
                errors.append("Source chunk is stale or superseded.")
            if action != "NO_CHANGE" and not row.reviewer_notes:
                errors.append("Reviewer notes are required.")
            if (
                action in {"APPROVE", "CORRECT_METADATA"}
                and row.retrieval_enabled
                and chunk.review_status == DocumentChunk.ReviewStatus.EXCLUDED
            ):
                errors.append(
                    "An excluded chunk cannot be retrieval-enabled in this action."
                )
            try:
                if action == "APPROVE":
                    if not row.chapter or not row.section:
                        raise ValidationError("APPROVE requires chapter and section.")
                    _validate_page_range(row, chunk)
                    if (
                        row.chapter != chunk.chapter
                        or row.section != chunk.section
                        or row.page_start != chunk.page_start
                        or row.page_end != chunk.page_end
                        or row.contains_warning != chunk.contains_warning
                        or row.contains_caution != chunk.contains_caution
                    ):
                        raise ValidationError(
                            "Use CORRECT_METADATA before approving changed metadata."
                        )
                elif action == "EXCLUDE" and row.retrieval_enabled:
                    raise ValidationError("An excluded chunk cannot enable retrieval.")
                elif action == "CORRECT_METADATA":
                    _validate_page_range(row, chunk)
                    if not _immutable_content_matches(row.content, chunk.content):
                        raise ValidationError(
                            "Metadata correction cannot change chunk content."
                        )
                    if (
                        re.search(r"\bwarning\b", chunk.content, re.IGNORECASE)
                        and not row.contains_warning
                    ):
                        raise ValidationError("WARNING text requires the warning flag.")
                    if (
                        re.search(r"\bcaution\b", chunk.content, re.IGNORECASE)
                        and not row.contains_caution
                    ):
                        raise ValidationError("CAUTION text requires the caution flag.")
                elif action == "SPLIT":
                    if chunk.origin != DocumentChunk.Origin.GENERATED:
                        raise ValidationError(
                            "Only a current generated chunk can be split."
                        )
                    segments = _split_segments(row)
                    validate_segments(
                        chunk,
                        segments,
                        safety_confirmed=_boolean(
                            row.correction_payload.get(
                                "safety_confirmed",
                                False,
                            ),
                            "safety_confirmed",
                        ),
                        safety_confirmation_required=(
                            chunk.contains_warning or chunk.contains_caution
                        ),
                        artifact_note=str(
                            row.correction_payload.get("artifact_note", "")
                        ),
                    )
                    for index, segment in enumerate(segments, 1):
                        if not segment.retrieval_enabled:
                            continue
                        child_hash = content_hash(segment.content)
                        key = (chunk.document_version_id, child_hash)
                        owner = active_hashes.get(key)
                        if owner:
                            raise ValidationError(
                                "Split child "
                                f"{index} duplicates retrieval-active chunk {owner}."
                            )
                        active_hashes[key] = f"{chunk.chunk_id} child {index}"
                if (
                    action in {"APPROVE", "CORRECT_METADATA"}
                    and (chunk.contains_warning or chunk.contains_caution)
                    and not _boolean(
                        row.correction_payload.get("safety_confirmed", False),
                        "safety_confirmed",
                    )
                ):
                    raise ValidationError(
                        "Safety confirmation is required for warning/caution content."
                    )
            except ValidationError as exc:
                errors.extend(exc.messages)

            will_enable = (
                action in {"APPROVE", "CORRECT_METADATA"} and row.retrieval_enabled
            )
            if action != "NO_CHANGE":
                key = (chunk.document_version_id, chunk.content_hash)
                if active_hashes.get(key) == chunk.chunk_id:
                    active_hashes.pop(key)
            if will_enable:
                key = (chunk.document_version_id, chunk.content_hash)
                owner = active_hashes.get(key)
                if owner and owner != chunk.chunk_id:
                    errors.append(f"Retrieval-active content duplicates chunk {owner}.")
                active_hashes[key] = chunk.chunk_id
        plans.append(
            RowPlan(
                row=row,
                chunk=chunk,
                errors=errors,
                summary=(
                    "No change" if action == "NO_CHANGE" else f"{action} {row.chunk_id}"
                ),
            )
        )
    return plans


def report_digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def dry_run_report(path: Path, plans: list[RowPlan]) -> dict[str, Any]:
    return {
        "source_file": str(path.resolve()),
        "source_sha256": report_digest(path),
        "valid": all(not plan.errors for plan in plans),
        "rows": [plan.as_report() for plan in plans],
    }


def metadata_payload(row: ReviewRow) -> dict[str, Any]:
    return {
        "chapter": row.chapter,
        "section": row.section,
        "page_start": row.page_start,
        "page_end": row.page_end,
        "contains_warning": row.contains_warning,
        "contains_caution": row.contains_caution,
        "retrieval_enabled": row.retrieval_enabled,
        "reviewer_notes": row.reviewer_notes,
    }


@transaction.atomic
def apply_review_plans(plans: list[RowPlan], reviewer: Any) -> list[dict[str, Any]]:
    if any(plan.errors for plan in plans):
        raise ValidationError("The review batch contains validation errors.")
    results: list[dict[str, Any]] = []
    now = timezone.now()
    for plan in plans:
        row = plan.row
        if row.proposed_action == "NO_CHANGE":
            results.append({"chunk_id": row.chunk_id, "result": "unchanged"})
            continue
        chunk = DocumentChunk.objects.select_for_update().get(pk=row.chunk_id)
        if chunk.content_hash != row.source_content_hash:
            raise ValidationError(f"{row.chunk_id}: source hash changed before apply.")
        if row.proposed_action == "SPLIT":
            correction = apply_split(
                chunk,
                _split_segments(row),
                reviewer=reviewer,
                artifact_note=str(row.correction_payload.get("artifact_note", "")),
                safety_confirmed=_boolean(
                    row.correction_payload.get("safety_confirmed", False),
                    "safety_confirmed",
                ),
                safety_confirmation_required=(
                    chunk.contains_warning or chunk.contains_caution
                ),
            )
            results.append(
                {
                    "chunk_id": row.chunk_id,
                    "result": "split",
                    "correction_id": str(correction.id),
                }
            )
            continue
        if row.proposed_action == "EXCLUDE":
            chunk.review_status = DocumentChunk.ReviewStatus.EXCLUDED
            chunk.retrieval_enabled = False
        elif row.proposed_action == "APPROVE":
            chunk.chapter = row.chapter
            chunk.section = row.section
            chunk.page_start = row.page_start
            chunk.page_end = row.page_end
            chunk.review_status = DocumentChunk.ReviewStatus.APPROVED
            chunk.retrieval_enabled = row.retrieval_enabled
        elif row.proposed_action == "CORRECT_METADATA":
            before = metadata_payload(
                ReviewRow(
                    **{
                        **asdict(row),
                        "chapter": chunk.chapter,
                        "section": chunk.section,
                        "page_start": chunk.page_start,
                        "page_end": chunk.page_end,
                        "contains_warning": chunk.contains_warning,
                        "contains_caution": chunk.contains_caution,
                        "retrieval_enabled": chunk.retrieval_enabled,
                        "reviewer_notes": chunk.reviewer_notes,
                    }
                )
            )
            after = metadata_payload(row)
            ChunkMetadataCorrection.objects.create(
                source_chunk=chunk,
                source_content_hash=chunk.content_hash,
                before_payload=before,
                after_payload=after,
                created_by=reviewer,
            )
            for field, value in after.items():
                setattr(chunk, field, value)
            chunk.review_status = DocumentChunk.ReviewStatus.APPROVED
        chunk.reviewer_notes = row.reviewer_notes
        chunk.reviewed_by = reviewer
        chunk.reviewed_at = now
        chunk.save()
        results.append({"chunk_id": row.chunk_id, "result": row.proposed_action})
    return results
